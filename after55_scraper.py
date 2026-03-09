"""
after55.com Active Adult Properties Scraper
============================================
Scrapes all 750 active adult properties from after55.com with full detail:
  - Property name, address, city, state, ZIP
  - Rent range, bed/bath options, sq footage
  - Unit count, stories, year built
  - Apartment features & community amenities
  - Hospital name + drive time/distance (up to 5 nearest)
  - Walk Score, Transit Score, Bike Score, Soundscore
  - Phone, office hours, lease terms
  - Application fee, pet policy
  - Listing URL

SETUP:
  pip install requests beautifulsoup4 openpyxl

RUN:
  python after55_scraper.py

OUTPUT:
  after55_active_adult_properties.xlsx  (created in same folder)
  after55_scraper.log                   (errors & progress)

The scraper respects the site with a 1–2 second delay between requests.
If you get rate-limited, increase DELAY_SECONDS below.
"""

import re
import time
import logging
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ────────────────────────────────────────────────────────────
DELAY_SECONDS = 1.5          # polite delay between requests
MAX_RETRIES   = 3            # retries on timeout/5xx
OUTPUT_FILE   = "after55_active_adult_properties.xlsx"
LOG_FILE      = "after55_scraper.log"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

BASE_URL      = "https://www.after55.com"
# Active adult search across US (all states)
SEARCH_STATES = [
    "al","ak","az","ar","ca","co","ct","de","fl","ga",
    "hi","id","il","in","ia","ks","ky","la","me","md",
    "ma","mi","mn","ms","mo","mt","ne","nv","nh","nj",
    "nm","ny","nc","nd","oh","ok","or","pa","ri","sc",
    "sd","tn","tx","ut","vt","va","wa","wv","wi","wy","dc",
]

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(),
    ]
)
log = logging.getLogger(__name__)

session = requests.Session()
session.headers.update(HEADERS)


def get(url, params=None):
    """GET with retries."""
    for attempt in range(MAX_RETRIES):
        try:
            r = session.get(url, params=params, timeout=15)
            r.raise_for_status()
            return r
        except requests.HTTPError as e:
            if r.status_code == 429:
                wait = 30 * (attempt + 1)
                log.warning(f"Rate limited. Waiting {wait}s …")
                time.sleep(wait)
            else:
                log.error(f"HTTP {r.status_code} for {url}")
                return None
        except Exception as e:
            log.warning(f"Attempt {attempt+1} failed for {url}: {e}")
            time.sleep(5)
    return None


def text(el):
    return el.get_text(strip=True) if el else ""


# ── Listing-page scraper ──────────────────────────────────────────────────────
def scrape_listing(url):
    """Scrape a single property detail page and return a dict."""
    r = get(url)
    if not r:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    data = {"listing_url": url}

    # ── Name & address ────────────────────────────────────────────────────
    h1 = soup.find("h1")
    data["name"] = text(h1)

    addr_el = soup.select_one("p.propertyAddress, [data-testid='property-address'], .property-address")
    if not addr_el:
        # fallback: look for address pattern near h1
        for tag in soup.find_all(["p", "div", "span"]):
            t = tag.get_text(strip=True)
            if re.search(r"\d{5}", t) and ("CA" in t or "TX" in t or len(t) < 80):
                addr_el = tag
                break
    full_addr = text(addr_el)
    # parse "2455 Colorado Blvd, Los Angeles, CA 90041"
    m = re.match(r"^(.*?),\s*(.*?),\s*([A-Z]{2})\s+(\d{5})", full_addr)
    if m:
        data["street"] = m.group(1).strip()
        data["city"]   = m.group(2).strip()
        data["state"]  = m.group(3).strip()
        data["zip"]    = m.group(4).strip()
    else:
        data["street"] = full_addr
        data["city"] = data["state"] = data["zip"] = ""

    # ── Rent / bed / bath / sqft summary ─────────────────────────────────
    summary_table = soup.find("table")
    if summary_table:
        cells = [text(td) for td in summary_table.find_all("td")]
        data["rent_range"] = cells[0] if len(cells) > 0 else ""
        data["bedrooms"]   = cells[1] if len(cells) > 1 else ""
        data["bathrooms"]  = cells[2] if len(cells) > 2 else ""
        data["sqft_range"] = cells[3] if len(cells) > 3 else ""
    else:
        data.update({"rent_range": "", "bedrooms": "", "bathrooms": "", "sqft_range": ""})

    # ── Property info (year built, units, stories) ────────────────────────
    prop_info = ""
    for tag in soup.find_all(["p", "li", "div"]):
        t = tag.get_text(strip=True)
        if "Built in" in t and "units" in t:
            prop_info = t
            break
    m_year  = re.search(r"Built in (\d{4})", prop_info)
    m_units = re.search(r"(\d+)\s+units", prop_info)
    m_stor  = re.search(r"(\d+)\s+stor", prop_info)
    data["year_built"] = m_year.group(1)  if m_year  else ""
    data["unit_count"] = m_units.group(1) if m_units else ""
    data["stories"]    = m_stor.group(1)  if m_stor  else ""

    # ── Lease term ────────────────────────────────────────────────────────
    lease_section = soup.find(string=re.compile("Lease Term"))
    data["lease_terms"] = ""
    if lease_section:
        parent = lease_section.find_parent()
        if parent:
            items = parent.find_next_siblings("li") or parent.parent.find_all("li")
            data["lease_terms"] = ", ".join(text(li) for li in items if text(li))

    # ── Fees ──────────────────────────────────────────────────────────────
    app_fee_match = re.search(r"Application Fee[^$]*\$(\d+)", r.text)
    data["application_fee"] = f"${app_fee_match.group(1)}" if app_fee_match else ""

    pet_match = re.search(r"(No Pets Allowed|Dogs Allowed|Cats Allowed|Pets Allowed)", r.text)
    data["pet_policy"] = pet_match.group(1) if pet_match else ""

    # ── Apartment features ────────────────────────────────────────────────
    apt_features = []
    apt_section = soup.find(string=re.compile("Apartment Features"))
    if apt_section:
        ul = apt_section.find_parent().find_next("ul")
        if ul:
            apt_features = [text(li) for li in ul.find_all("li")]
    data["apartment_features"] = " | ".join(apt_features)

    # ── Community features ────────────────────────────────────────────────
    comm_features = []
    comm_section = soup.find(string=re.compile("Community Features"))
    if comm_section:
        ul = comm_section.find_parent().find_next("ul")
        if ul:
            comm_features = [text(li) for li in ul.find_all("li")]
    data["community_features"] = " | ".join(comm_features)

    # ── Hospitals ─────────────────────────────────────────────────────────
    hospitals = []
    hosp_section = soup.find(string=re.compile("Hospitals"))
    if hosp_section:
        table = hosp_section.find_parent().find_next("table")
        if table:
            for row in table.find_all("tr")[1:]:  # skip header
                cols = row.find_all("td")
                if len(cols) >= 2:
                    hosp_name = text(cols[0])
                    hosp_dist = text(cols[1])
                    hospitals.append(f"{hosp_name} ({hosp_dist})")
    for i in range(5):
        data[f"hospital_{i+1}"] = hospitals[i] if i < len(hospitals) else ""

    # ── Scores ────────────────────────────────────────────────────────────
    scores = {}
    for label in ["Walk Score", "Transit Score", "Bike Score", "Soundscore"]:
        m = re.search(rf"{label}[^0-9]*(\d+)\s*/\s*100", r.text)
        scores[label] = m.group(1) if m else ""
    data["walk_score"]    = scores["Walk Score"]
    data["transit_score"] = scores["Transit Score"]
    data["bike_score"]    = scores["Bike Score"]
    data["sound_score"]   = scores["Soundscore"]

    # ── Phone ─────────────────────────────────────────────────────────────
    phone_match = re.search(r'tel:\+1(\d{10})', r.text)
    if phone_match:
        p = phone_match.group(1)
        data["phone"] = f"({p[:3]}) {p[3:6]}-{p[6:]}"
    else:
        data["phone"] = ""

    # ── Office hours ──────────────────────────────────────────────────────
    hours = []
    for day in ["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]:
        m = re.search(rf"{day}\s*(.+?)(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday|$)",
                      r.text, re.DOTALL)
        if m:
            h = m.group(1).strip().split("\n")[0].strip()
            if h and len(h) < 50:
                hours.append(f"{day[:3]}: {h}")
    data["office_hours"] = " | ".join(hours)

    return data


# ── Search-page scraper ───────────────────────────────────────────────────────
def get_listing_urls_for_state(state_code):
    """Collect all listing URLs for a given state's active adult search."""
    urls = []
    page = 1
    while True:
        if page == 1:
            url = f"{BASE_URL}/search/{state_code}/specialties-active-adult"
        else:
            url = f"{BASE_URL}/search/{state_code}/specialties-active-adult/page-{page}"

        r = get(url)
        if not r:
            break
        soup = BeautifulSoup(r.text, "html.parser")

        # find all property links
        links = soup.select("a[href*='/ca/'], a[href*='/tx/'], a[href*='/fl/']")
        # generic: any link with a property-style path /{state}/{city}/{slug}/{id}
        links = soup.select("a[href]")
        new_urls = []
        for a in links:
            href = a.get("href", "")
            # property pages look like /ca/los-angeles/property-name/abc123
            if re.match(r"^/[a-z]{2}/[^/]+/[^/]+/[a-z0-9]{7}$", href):
                full = BASE_URL + href
                if full not in urls and full not in new_urls:
                    new_urls.append(full)

        if not new_urls:
            break

        urls.extend(new_urls)
        log.info(f"  [{state_code.upper()}] Page {page}: found {len(new_urls)} listings (total {len(urls)})")

        # check if there's a next page
        next_link = soup.find("a", string=re.compile(r"page \d+", re.I))
        pagination = soup.select("a[href*='page-']")
        current_page_nums = [
            int(re.search(r"page-(\d+)", a["href"]).group(1))
            for a in pagination if re.search(r"page-(\d+)", a.get("href",""))
        ]
        if not current_page_nums or page >= max(current_page_nums):
            break

        page += 1
        time.sleep(DELAY_SECONDS)

    return urls


# ── Excel writer ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("listing_url",          "Listing URL"),
    ("name",                 "Property Name"),
    ("street",               "Street Address"),
    ("city",                 "City"),
    ("state",                "State"),
    ("zip",                  "ZIP"),
    ("phone",                "Phone"),
    ("rent_range",           "Rent Range"),
    ("bedrooms",             "Bedrooms"),
    ("bathrooms",            "Bathrooms"),
    ("sqft_range",           "Sq Ft Range"),
    ("year_built",           "Year Built"),
    ("unit_count",           "Unit Count"),
    ("stories",              "Stories"),
    ("lease_terms",          "Lease Terms"),
    ("application_fee",      "Application Fee"),
    ("pet_policy",           "Pet Policy"),
    ("apartment_features",   "Apartment Features"),
    ("community_features",   "Community Features"),
    ("hospital_1",           "Nearest Hospital"),
    ("hospital_2",           "Hospital 2"),
    ("hospital_3",           "Hospital 3"),
    ("hospital_4",           "Hospital 4"),
    ("hospital_5",           "Hospital 5"),
    ("walk_score",           "Walk Score"),
    ("transit_score",        "Transit Score"),
    ("bike_score",           "Bike Score"),
    ("sound_score",          "Sound Score"),
    ("office_hours",         "Office Hours"),
]

def write_excel(rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active Adult Properties"

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    data_font = Font(name="Arial", size=10)
    url_font  = Font(name="Arial", size=10, color="0563C1", underline="single")

    for col, (_, header) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_i, prop in enumerate(rows, 2):
        fill = alt_fill if row_i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col, (key, _) in enumerate(COLUMNS, 1):
            val = prop.get(key, "")
            c = ws.cell(row=row_i, column=col, value=val)
            c.fill = fill
            c.font = url_font if key == "listing_url" else data_font
            c.alignment = Alignment(vertical="center", wrap_text=False)

    # column widths
    widths = {
        "listing_url": 55, "name": 38, "street": 28, "city": 16,
        "state": 7, "zip": 8, "phone": 16, "rent_range": 20,
        "bedrooms": 14, "bathrooms": 12, "sqft_range": 14,
        "year_built": 10, "unit_count": 10, "stories": 8,
        "lease_terms": 16, "application_fee": 14, "pet_policy": 16,
        "apartment_features": 55, "community_features": 45,
        "hospital_1": 40, "hospital_2": 40, "hospital_3": 40,
        "hospital_4": 40, "hospital_5": 40,
        "walk_score": 10, "transit_score": 12, "bike_score": 10,
        "sound_score": 11, "office_hours": 50,
    }
    for col, (key, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Source",            "after55.com/search/specialties-active-adult"),
        ("Date Scraped",      time.strftime("%Y-%m-%d")),
        ("Total Properties",  len(rows)),
        ("Fields Captured",   len(COLUMNS)),
    ]
    for r, (k, v) in enumerate(summary, 1):
        ws2.cell(row=r, column=1, value=k).font  = Font(name="Arial", bold=True)
        ws2.cell(row=r, column=2, value=v).font  = Font(name="Arial")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 55

    wb.save(filename)
    log.info(f"Saved {len(rows)} properties → {filename}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    all_listing_urls = []

    log.info("=== Phase 1: Collecting listing URLs from all states ===")
    for state in SEARCH_STATES:
        state_urls = get_listing_urls_for_state(state)
        all_listing_urls.extend(state_urls)
        log.info(f"[{state.upper()}] {len(state_urls)} listings found. Running total: {len(all_listing_urls)}")
        time.sleep(DELAY_SECONDS)

    # deduplicate
    all_listing_urls = list(dict.fromkeys(all_listing_urls))
    log.info(f"\n=== Phase 2: Scraping {len(all_listing_urls)} unique property pages ===")

    results = []
    for i, url in enumerate(all_listing_urls, 1):
        log.info(f"[{i}/{len(all_listing_urls)}] {url}")
        prop = scrape_listing(url)
        if prop:
            results.append(prop)
        else:
            log.warning(f"  Skipped (no data returned)")
        time.sleep(DELAY_SECONDS)

        # checkpoint save every 50 properties
        if i % 50 == 0:
            write_excel(results, OUTPUT_FILE)
            log.info(f"  Checkpoint saved ({len(results)} properties so far)")

    write_excel(results, OUTPUT_FILE)
    log.info(f"\nDone. {len(results)} properties written to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
